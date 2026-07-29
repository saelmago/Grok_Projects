from openpyxl import load_workbook
from openpyxl.styles import Font
from pathlib import Path

root = Path(__file__).resolve().parent.parent
path = root / "inventario" / "Arsenal_Multioficio_2026.xlsx"
alt = root / "inventario" / "Arsenal_Multioficio_2026_v2.xlsx"

link_font = Font(name="Arial", size=10, color="0563C1", underline="single")
url_font = Font(name="Arial", size=8, color="6B7280")

klein_urls = {
    "NCVT-39": "https://www.kleintools.com/catalog/electrical-testers/dual-range-non-contact-voltage-tester-flashlight-12-1000v-ac",
    "RT250": "https://www.kleintools.com/catalog/electrical-testers/gfci-receptacle-tester-lcd",
    "ET310": "https://www.kleintools.com/catalog/electrical-testers/digital-circuit-breaker-finder-gfci-outlet-tester",
    "VDV500-705P": "https://www.kleintools.com/catalog/tone-probe/tone-probe-test-and-trace-kit",
    "32306INS": "https://www.kleintools.com/catalog/electricians-screwdrivers/6-1-insulated-auto-lock-screwdriver",
    "2139NERINS": "https://www.kleintools.com/catalog/electricians-pliers/insulated-pliers-side-cutters-9-inch",
    "2288RINS": "https://www.kleintools.com/catalog/electricians-pliers/diagonal-cutting-pliers-insulated-high-leverage-8-inch",
    "2038RINS": "https://www.kleintools.com/catalog/insulated-pliers/pliers-long-nose-side-cutters-insulated-8-inch",
    "11055RINS": "https://www.kleintools.com/catalog/electricians-cutting-and-crimping-tools/insulated-klein-kurve-wire-stripper-and-cutter",
    "32288": "https://www.kleintools.com/catalog/electricians-tool-sets/8-1-insulated-interchangeable-screwdriver-set",
    "1005RINS": "https://www.kleintools.com/catalog/electricians-cutting-and-crimping-tools/crimping-and-cutting-tool-connectors-insulated",
    "32751": "https://www.kleintools.com/catalog/multi-bit-screwdrivers/adjustable-screwdriver-2-phillips-14-inch-slotted",
    "33809M": "https://www.kleintools.com/catalog/impact-socket-sets/grab-and-go-impact-socket-set-metric-10-piece",
    "65400": "https://www.kleintools.com/catalog/socket-wrenches/knect-38-inch-drive-impact-rated-pass-through-socket-set-15-piece",
    "CL810": "https://www.kleintools.com/catalog/clamp-meters/600a-acdc-auto-ranging-trms-clamp-meter-worklight",
}

wb = load_workbook(path)
ws = wb["Inventario"]
wp = wb["Pendientes"]
updated = []

for row in range(5, ws.max_row + 1):
    if ws.cell(row, 2).value != "Klein":
        continue
    modelo = str(ws.cell(row, 6).value or "").strip()
    url = klein_urls.get(modelo)
    if not url:
        print("MISSING INV", modelo)
        continue
    label = ws.cell(row, 8)
    label.value = "Ver ficha Klein"
    label.font = link_font
    label.hyperlink = url
    url_cell = ws.cell(row, 9)
    url_cell.value = url
    url_cell.font = url_font
    url_cell.hyperlink = url
    updated.append(("Inv", modelo))

for row in range(5, wp.max_row + 1):
    if wp.cell(row, 2).value != "Klein":
        continue
    modelo = str(wp.cell(row, 4).value or "").strip()
    url = klein_urls.get(modelo)
    if not url:
        print("MISSING PEND", modelo)
        continue
    label = wp.cell(row, 7)
    label.value = "Ver ficha Klein"
    label.font = link_font
    label.hyperlink = url
    url_cell = wp.cell(row, 8)
    url_cell.value = url
    url_cell.font = url_font
    url_cell.hyperlink = url
    updated.append(("Pend", modelo))

ws["A3"] = (
    "Enlaces Klein: fichas oficiales del catalogo. "
    "NCVT-39 apunta a NCVT3P (modelo actual). VDV500-705P apunta al kit VDV500-705."
)
ws["A3"].font = Font(name="Arial", size=9, color="6B7280", italic=True)

try:
    wb.save(path)
    print("Saved", path)
except PermissionError:
    wb.save(alt)
    print("File locked; saved", alt)

print("Updated", len(updated), "Klein links")
for kind, model in updated:
    print(kind, model)
