"""
Genera inventario/Arsenal_Multioficio_2026.xlsx

Codigo por linea de inventario (inmutable, no por unidad fisica):
  {MARCA}-{FAMILIA}-{###}
  Ej: MW-M12-001, KL-ELC-003, MW-M12-BAT-001

Columnas Inventario:
  #, Codigo, Marca, Tipo, Categoria, Herramienta, Modelo, Cantidad, Estado, Enlace marca, URL
"""
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# (codigo, marca, tipo, cat, nombre, modelo, cantidad, estado, url)
INVENTARIO = [
    ("MW-FTN-001", "Milwaukee", "Herramienta", "Fontaneria", "M12 Stick Transfer Pump", "2579-20", 1, "Comprado",
     "https://www.milwaukeetool.com/Products/Power-Tools/Plumbing/Pumps/M12-FUEL-Stick-Transfer-Pump/2579-20"),
    ("MW-FTN-002", "Milwaukee", "Herramienta", "Fontaneria", "Llave Faucet Swap-Out", "48-22-7100", 1, "Ya tengo",
     "https://www.milwaukeetool.com/Products/Hand-Tools/Plumbing/Wrenches/Faucet-and-Sink-Installer/48-22-7100"),
    ("KP-FTN-001", "Knipex", "Herramienta", "Fontaneria", 'Cobra Water Pump Pliers 10"', "87 02 250", 1, "Ya tengo",
     "https://www.knipex.com/products/water-pump-pliers-and-pipe-wrenches/water-pump-pliers/knipex-cobra-water-pump-pliers/8702250"),
    ("KP-FTN-002", "Knipex", "Herramienta", "Fontaneria", 'TwinGrip Pliers 8"', "83 01 200", 1, "Ya tengo",
     "https://www.knipex.com/products/pliers/front-and-side-gripping-pliers/knipex-twingrip-front-and-side-gripping-pliers/8201200"),
    ("KP-FTN-003", "Knipex", "Herramienta", "Fontaneria", 'Pliers Wrench 10"', "86 01 250", 1, "Comprado",
     "https://www.knipex.com/products/pliers-wrenches-and-pipe-wrenches/pliers-wrenches/pliers-wrench-pliers-and-a-wrench-in-a-single-tool/8601250"),
    ("WR-FTN-001", "Wera", "Herramienta", "Fontaneria", "Joker Self-Setting Wrench M/L/XL", "6004 serie", 1, "Ya tengo",
     "https://www.wera.de/en/tools/tool-types/wrench/joker-adjustable-open-end-wrenches"),
    ("KL-ELC-001", "Klein", "Herramienta", "Electricidad", "Non-Contact Voltage Tester", "NCVT-39", 1, "Ya tengo",
     "https://www.kleintools.com/catalog/electrical-testers/dual-range-non-contact-voltage-tester-flashlight-12-1000v-ac"),
    ("KL-ELC-002", "Klein", "Herramienta", "Electricidad", "GFCI Receptacle Tester", "RT250", 1, "Ya tengo",
     "https://www.kleintools.com/catalog/electrical-testers/gfci-receptacle-tester-lcd"),
    ("KL-ELC-003", "Klein", "Herramienta", "Electricidad", "Circuit Breaker Finder", "ET310", 1, "Ya tengo",
     "https://www.kleintools.com/catalog/electrical-testers/digital-circuit-breaker-finder-gfci-outlet-tester"),
    ("KL-ELC-004", "Klein", "Herramienta", "Electricidad", "Tone and Probe Kit", "VDV500-705P", 1, "Ya tengo",
     "https://www.kleintools.com/catalog/tone-probe/tone-probe-test-and-trace-kit"),
    ("KL-ELC-005", "Klein", "Herramienta", "Electricidad", "6-in-1 Insulated Screwdriver", "32306INS", 1, "Ya tengo",
     "https://www.kleintools.com/catalog/electricians-screwdrivers/6-1-insulated-auto-lock-screwdriver"),
    ("KL-ELC-006", "Klein", "Herramienta", "Electricidad", '9" Insulated Side-Cutting Pliers', "2139NERINS", 1, "Ya tengo",
     "https://www.kleintools.com/catalog/electricians-pliers/insulated-pliers-side-cutters-9-inch"),
    ("KL-ELC-007", "Klein", "Herramienta", "Electricidad", '8" Insulated Diagonal Cutting Pliers', "2288RINS", 1, "Ya tengo",
     "https://www.kleintools.com/catalog/electricians-pliers/diagonal-cutting-pliers-insulated-high-leverage-8-inch"),
    ("KL-ELC-008", "Klein", "Herramienta", "Electricidad", '8" Insulated Long Nose Pliers', "2038RINS", 1, "Ya tengo",
     "https://www.kleintools.com/catalog/insulated-pliers/pliers-long-nose-side-cutters-insulated-8-inch"),
    ("KL-ELC-009", "Klein", "Herramienta", "Electricidad", "Insulated Wire Stripper", "11055RINS", 1, "Ya tengo",
     "https://www.kleintools.com/catalog/electricians-cutting-and-crimping-tools/insulated-klein-kurve-wire-stripper-and-cutter"),
    ("KL-ELC-010", "Klein", "Herramienta", "Electricidad", "8-in-1 Insulated Screwdriver Set", "32288", 1, "Ya tengo",
     "https://www.kleintools.com/catalog/electricians-tool-sets/8-1-insulated-interchangeable-screwdriver-set"),
    ("KL-ELC-011", "Klein", "Herramienta", "Electricidad", "Insulated Crimping and Cutting Tool", "1005RINS", 1, "Comprado",
     "https://www.kleintools.com/catalog/electricians-cutting-and-crimping-tools/crimping-and-cutting-tool-connectors-insulated"),
    ("KL-ELC-012", "Klein", "Herramienta", "Electricidad", "Adjustable Length Screwdriver", "32751", 1, "Comprado",
     "https://www.kleintools.com/catalog/multi-bit-screwdrivers/adjustable-screwdriver-2-phillips-14-inch-slotted"),
    ("KL-ELC-013", "Klein", "Accesorio", "Electricidad", "Grab-And-Go Impact Socket Set (metrico)", "33809M", 1, "Comprado",
     "https://www.kleintools.com/catalog/impact-socket-sets/grab-and-go-impact-socket-set-metric-10-piece"),
    ("MW-M12-001", "Milwaukee", "Herramienta", "Electrica M12", "FUEL Installation Drill/Driver 4-en-1", "2505-20", 1, "Ya tengo",
     "https://www.milwaukeetool.com/Products/Power-Tools/Drilling/Installation-Drill-Drivers/M12-FUEL-Installation-Drill-Driver/2505-20"),
    ("MW-M12-002", "Milwaukee", "Herramienta", "Electrica M12", "FUEL Stubby Impact Wrench", "2563-20", 1, "Ya tengo",
     "https://www.milwaukeetool.com/Search?q=2563-20"),
    ("MW-M12-003", "Milwaukee", "Herramienta", "Electrica M12", "FUEL Hammer Drill", "3404-20", 1, "Ya tengo",
     "https://www.milwaukeetool.com/Search?q=3404-20"),
    ("MW-M12-004", "Milwaukee", "Herramienta", "Electrica M12", "FUEL Impact Driver", "3453-20", 1, "Ya tengo",
     "https://www.milwaukeetool.com/Search?q=3453-20"),
    ("MW-M12-005", "Milwaukee", "Herramienta", "Electrica M12", "FUEL Oscillating Multi-Tool", "2526-20", 1, "Ya tengo",
     "https://www.milwaukeetool.com/Search?q=2526-20"),
    ("MW-M12-006", "Milwaukee", "Herramienta", "Electrica M12", "ROVER Flood Light", "2367-20", 1, "Ya tengo",
     "https://www.milwaukeetool.com/Search?q=2367-20"),
    ("MW-M12-007", "Milwaukee", "Herramienta", "Electrica M12", 'FUEL 3" Compact Cut Off Tool', "2522-20", 1, "Comprado",
     "https://www.milwaukeetool.com/Search?q=2522-20"),
    ("MW-M18-001", "Milwaukee", "Herramienta", "Electrica M18", "M18 FUEL Sawzall (Sierra Sable)", "2719-20", 1, "Comprado",
     "https://www.milwaukeetool.com/Search?q=2719-20"),
    ("MW-M12-BAT-001", "Milwaukee", "Accesorio", "Electrica M12", "Bateria M12 CP 2.0 Ah", "CP 2.0", 3, "Ya tengo",
     "https://www.milwaukeetool.com/Products/Power-Tools/Batteries-Chargers/Batteries"),
    ("MW-M12-BAT-002", "Milwaukee", "Accesorio", "Electrica M12", "Bateria M12 XC 4.0 Ah", "XC 4.0", 1, "Ya tengo",
     "https://www.milwaukeetool.com/Products/Power-Tools/Batteries-Chargers/Batteries"),
    ("MW-M18-BAT-001", "Milwaukee", "Accesorio", "Electrica M18", "Bateria M18 XC 5.0 Ah", "XC 5.0", 3, "Ya tengo",
     "https://www.milwaukeetool.com/Products/Power-Tools/Batteries-Chargers/Batteries"),
    ("MW-M18-CHG-001", "Milwaukee", "Accesorio", "Electrica M18", "Cargador 12-18V", "12-18V", 1, "Comprado",
     "https://www.milwaukeetool.com/Products/Power-Tools/Batteries-Chargers/Chargers"),
    ("MW-PKO-001", "Milwaukee", "Organizacion", "Organizacion", "PACKOUT Backpack", "48-22-8301", 1, "Comprado",
     "https://www.milwaukeetool.com/Search?q=48-22-8301"),
    ("MW-PKO-002", "Milwaukee", "Organizacion", "Organizacion", "PACKOUT Compact Organizer", "48-22-8435", 1, "Comprado",
     "https://www.milwaukeetool.com/Search?q=48-22-8435"),
    ("MW-PKO-005", "Milwaukee", "Organizacion", "Organizacion", "PACKOUT Organizer 19 pulg", "48-22-8431", 1, "Comprado",
     "https://www.milwaukeetool.com/Search?q=48-22-8431"),
    ("SK-ELC-001", "Skil", "Herramienta", "Electrica", "Martillo Automatico", "AH6552A-00", 1, "Ya tengo",
     "https://www.skil.com/search?q=AH6552A"),
    ("AN-ELC-001", "Antive", "Herramienta", "Electrica", "Tijeras Electricas Inalambricas", "E-S01", 1, "Ya tengo",
     "https://www.amazon.com/s?k=Antive+E-S01+electric+scissors"),
    ("HP-PRE-001", "Huepar", "Herramienta", "Precision", "Nivel Laser Verde Cross-Line", "BOX-1G", 1, "Ya tengo",
     "https://huepar.com/products/huepar-box1g-laser-level"),
    ("FR-PRE-001", "Franklin", "Herramienta", "Precision", "Nivel Digital Electronico", "iA12", 1, "Ya tengo",
     "https://www.amazon.com/s?k=Franklin+Sensors+iA12"),
    ("LX-SLD-001", "LEXIVON", "Herramienta", "Soldadura", "Soldador de Butano 7 puntas", "LX-770", 1, "Ya tengo",
     "https://www.lexivon.com/products/lx-770"),
    ("MW-SEG-001", "Milwaukee", "Seguridad", "Seguridad", "Lentes de seguridad", "48-73-2013", 1, "Ya tengo",
     "https://www.milwaukeetool.com/Search?q=48-73-2013"),
    ("MW-SEG-002", "Milwaukee", "Seguridad", "Seguridad", "Guantes PACKOUT Cut Level 1 Smartswipe", "48-73-2013", 1, "Ya tengo",
     "https://www.milwaukeetool.com/Search?q=PACKOUT+Cut+Level+Smartswipe+gloves"),
    ("MW-ACC-001", "Milwaukee", "Accesorio", "Accesorio", "Juego de 3 hojas para Multi-Tool", "49-10-9001", 1, "Comprado",
     "https://www.milwaukeetool.com/Search?q=49-10-9001"),
]

# (codigo_reservado, prio, marca, nombre, modelo, precio, notas, url)
PENDIENTES = [
    ("KL-SOC-001", 1, "Klein", "KNECT 15-Piece Pass Through Socket Set", "65400", None, "Sockets",
     "https://www.kleintools.com/catalog/socket-wrenches/knect-38-inch-drive-impact-rated-pass-through-socket-set-15-piece"),
    ("KL-MED-001", 2, "Klein", "Clamp Meter 600A", "CL810", 80000, "Medicion",
     "https://www.kleintools.com/catalog/clamp-meters/600a-acdc-auto-ranging-trms-clamp-meter-worklight"),
    ("MW-PKO-003", 3, "Milwaukee", "PACKOUT Rolling Drawer Tool Box", "48-22-8420", None, "Organizacion",
     "https://www.milwaukeetool.com/Search?q=48-22-8420"),
    ("MW-PKO-004", 4, "Milwaukee", "PACKOUT 2-Wheel Utility Cart", "48-22-8415", None, "Organizacion",
     "https://www.milwaukeetool.com/Search?q=48-22-8415"),
    ("MW-M12-008", 5, "Milwaukee", "M12 AIRSNAKE", "2572B-21", 280000, "Fontaneria",
     "https://www.milwaukeetool.com/Search?q=2572B-21"),
    ("CNS-ORD-AMZ-001", 6, "Amazon", "Terminar de cargar orden consumibles Amazon", "orden haisstronica", None, "Consumibles",
     "https://www.amazon.com/"),
]

# Consumibles: por KIT/paquete (no pieza a pieza).
# (codigo, clase, tipo_terminal, nombre, marca_ref, forma, cantidad, umbral_reorden, estado, ubicacion, notas, url)
# ubicacion = codigo del contenedor/lugar en inventario (ej. MW-PKO-005 = PACKOUT Organizer 19")
# clase: Conector electrico | Cinta / aislante | Fijacion | Corte / desgaste | Quimico / sellado | Otro
# forma: kit | paquete | aprox
# estado: Tengo | Reponer
# tipo_terminal: solo si clase = Conector electrico; si no, "—"
# Spade por COLOR. Kit haisstronica 160PCS (Amazon B0D1K6L6TL) -> ubicacion MW-PKO-005
CLASES_CONSUMIBLE = [
    "Conector electrico",
    "Cinta / aislante",
    "Fijacion",
    "Corte / desgaste",
    "Quimico / sellado",
    "Otro",
]
# Codigos de ubicacion conocidos (inventario organizacion)
UBIC_PKO_ORG19 = "MW-PKO-005"  # PACKOUT Organizer 19" 48-22-8431

CONSUMIBLES = [
    (
        "CNS-SPD-ROJ-001",
        "Conector electrico",
        "Spade (pala / quick disconnect)",
        "Heat shrink Male/Female — ROJO 22-16 AWG",
        "haisstronica B0D1K6L6TL",
        "aprox",
        60,
        15,
        "Tengo",
        UBIC_PKO_ORG19,
        "En PACKOUT Organizer 19 (MW-PKO-005). SPADE rojo. Kit 160PCS. Aprox al reponer.",
        "https://www.amazon.com/dp/B0D1K6L6TL",
    ),
    (
        "CNS-SPD-AZU-001",
        "Conector electrico",
        "Spade (pala / quick disconnect)",
        "Heat shrink Male/Female — AZUL 16-14 AWG",
        "haisstronica B0D1K6L6TL",
        "aprox",
        60,
        15,
        "Tengo",
        UBIC_PKO_ORG19,
        "En PACKOUT Organizer 19 (MW-PKO-005). SPADE azul. Kit 160PCS. Aprox al reponer.",
        "https://www.amazon.com/dp/B0D1K6L6TL",
    ),
    (
        "CNS-SPD-AMA-001",
        "Conector electrico",
        "Spade (pala / quick disconnect)",
        "Heat shrink Male/Female — AMARILLO 12-10 AWG",
        "haisstronica B0D1K6L6TL",
        "aprox",
        40,
        10,
        "Tengo",
        UBIC_PKO_ORG19,
        "En PACKOUT Organizer 19 (MW-PKO-005). SPADE amarillo. Kit 160PCS. Aprox al reponer.",
        "https://www.amazon.com/dp/B0D1K6L6TL",
    ),
    # Butt / Solder seal 120PCS — haisstronica B07C3NBTJ9 — cantidades oficiales del packaging:
    # Blanco 26-24: 25 | Rojo 22-18: 45 | Azul 16-14: 40 | Amarillo 12-10: 10  = 120
    (
        "CNS-BTT-BLA-001",
        "Conector electrico",
        "Butt (empalme / solder seal)",
        "Solder seal butt — BLANCO 26-24 AWG",
        "haisstronica B07C3NBTJ9",
        "aprox",
        25,
        6,
        "Tengo",
        UBIC_PKO_ORG19,
        "Butt solder seal. BLANCO 26-24 (0.25-0.34mm2) x25. Kit 120PCS. Ubic. MW-PKO-005.",
        "https://www.amazon.com/dp/B07C3NBTJ9",
    ),
    (
        "CNS-BTT-ROJ-001",
        "Conector electrico",
        "Butt (empalme / solder seal)",
        "Solder seal butt — ROJO 22-18 AWG",
        "haisstronica B07C3NBTJ9",
        "aprox",
        45,
        10,
        "Tengo",
        UBIC_PKO_ORG19,
        "Butt solder seal. ROJO 22-18 (0.5-1.0mm2) x45. Kit 120PCS. Ubic. MW-PKO-005.",
        "https://www.amazon.com/dp/B07C3NBTJ9",
    ),
    (
        "CNS-BTT-AZU-001",
        "Conector electrico",
        "Butt (empalme / solder seal)",
        "Solder seal butt — AZUL 16-14 AWG",
        "haisstronica B07C3NBTJ9",
        "aprox",
        40,
        10,
        "Tengo",
        UBIC_PKO_ORG19,
        "Butt solder seal. AZUL 16-14 (1.5-2.5mm2) x40. Kit 120PCS. Ubic. MW-PKO-005.",
        "https://www.amazon.com/dp/B07C3NBTJ9",
    ),
    (
        "CNS-BTT-AMA-001",
        "Conector electrico",
        "Butt (empalme / solder seal)",
        "Solder seal butt — AMARILLO 12-10 AWG",
        "haisstronica B07C3NBTJ9",
        "aprox",
        10,
        3,
        "Tengo",
        UBIC_PKO_ORG19,
        "Butt solder seal. AMARILLO 12-10 (4.0-6.0mm2) x10. Kit 120PCS. Ubic. MW-PKO-005.",
        "https://www.amazon.com/dp/B07C3NBTJ9",
    ),
    # Heat shrink CRIMP butt 200PCS — haisstronica B07RX6QYX5 — modelos BHT-*; NO solder seal
    # Blanco BHT-0.5 26-24 | Rosa BHT-1.25 22-16 | Azul BHT-2 16-14 | Amarillo BHT-5.5 12-10
    # Cant. 50 c/u = 200 (repartir igual; ajustar si el blister dice otra cosa)
    (
        "CNS-BHC-BLA-001",
        "Conector electrico",
        "Butt (empalme / heat shrink crimp)",
        "Heat shrink crimp butt BHT-0.5 — BLANCO 26-24 AWG (0.5mm)",
        "haisstronica B07RX6QYX5 | BHT-0.5",
        "aprox",
        50,
        12,
        "Tengo",
        UBIC_PKO_ORG19,
        "Butt heat-shrink CRIMP (no solder). BHT-0.5 blanco. Kit 200PCS. Ubic. MW-PKO-005.",
        "https://www.amazon.com/dp/B07RX6QYX5",
    ),
    (
        "CNS-BHC-ROS-001",
        "Conector electrico",
        "Butt (empalme / heat shrink crimp)",
        "Heat shrink crimp butt BHT-1.25 — ROSA 22-16 AWG (0.7mm)",
        "haisstronica B07RX6QYX5 | BHT-1.25",
        "aprox",
        50,
        12,
        "Tengo",
        UBIC_PKO_ORG19,
        "Butt heat-shrink CRIMP. BHT-1.25 rosa/pink (a veces listado rojo). Kit 200PCS. Ubic. MW-PKO-005.",
        "https://www.amazon.com/dp/B07RX6QYX5",
    ),
    (
        "CNS-BHC-AZU-001",
        "Conector electrico",
        "Butt (empalme / heat shrink crimp)",
        "Heat shrink crimp butt BHT-2 — AZUL 16-14 AWG (0.8mm)",
        "haisstronica B07RX6QYX5 | BHT-2",
        "aprox",
        50,
        12,
        "Tengo",
        UBIC_PKO_ORG19,
        "Butt heat-shrink CRIMP. BHT-2 azul. Kit 200PCS. Ubic. MW-PKO-005.",
        "https://www.amazon.com/dp/B07RX6QYX5",
    ),
    (
        "CNS-BHC-AMA-001",
        "Conector electrico",
        "Butt (empalme / heat shrink crimp)",
        "Heat shrink crimp butt BHT-5.5 — AMARILLO 12-10 AWG (1.0mm)",
        "haisstronica B07RX6QYX5 | BHT-5.5",
        "aprox",
        50,
        12,
        "Tengo",
        UBIC_PKO_ORG19,
        "Butt heat-shrink CRIMP. BHT-5.5 amarillo. Kit 200PCS. Ubic. MW-PKO-005.",
        "https://www.amazon.com/dp/B07RX6QYX5",
    ),
    # Non-insulated butt 50PCS AWG 12-10 tinned copper — haisstronica B0F1N2WH51
    (
        "CNS-BNI-1210-001",
        "Conector electrico",
        "Butt (empalme / non-insulated crimp)",
        "Non-insulated butt tinned copper — AWG 12-10 (50PCS)",
        "haisstronica B0F1N2WH51",
        "paquete",
        50,
        12,
        "Tengo",
        UBIC_PKO_ORG19,
        "Butt NO aislado, cobre estanado. Solo 12-10 AWG x50. Recomienda heat shrink aparte. Ubic. MW-PKO-005.",
        "https://www.amazon.com/dp/B0F1N2WH51",
    ),
]

wb = Workbook()

header_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
title_font = Font(name="Arial", bold=True, size=16, color="1F2937")
subtitle_font = Font(name="Arial", size=10, color="4B5563")
body_font = Font(name="Arial", size=10)
link_font = Font(name="Arial", size=10, color="0563C1", underline="single")
header_fill = PatternFill("solid", fgColor="1E3A5F")
alt_fill = PatternFill("solid", fgColor="F3F4F6")
comprado_fill = PatternFill("solid", fgColor="DCFCE7")
tengo_fill = PatternFill("solid", fgColor="DBEAFE")
pend_fill = PatternFill("solid", fgColor="FEF3C7")
cns_fill = PatternFill("solid", fgColor="E0E7FF")
cns_header = PatternFill("solid", fgColor="4338CA")
reponer_fill = PatternFill("solid", fgColor="FEE2E2")
tengo_cns_fill = PatternFill("solid", fgColor="D1FAE5")
resumen_fill = PatternFill("solid", fgColor="EEF2FF")
code_fill = PatternFill("solid", fgColor="FEF9C3")
thin = Border(
    left=Side(style="thin", color="D1D5DB"),
    right=Side(style="thin", color="D1D5DB"),
    top=Side(style="thin", color="D1D5DB"),
    bottom=Side(style="thin", color="D1D5DB"),
)
center = Alignment(horizontal="center", vertical="center", wrap_text=True)
left = Alignment(horizontal="left", vertical="center", wrap_text=True)

# ===== Inventario =====
ws = wb.active
ws.title = "Inventario"
ws["A1"] = "Inventario - Arsenal Multioficio 2026"
ws["A1"].font = title_font
ws.merge_cells("A1:K1")
ws["A2"] = "Propietario: Saul Marin Gonzalez"
ws["A2"].font = subtitle_font
ws["C2"] = "Ultima actualizacion: 28/07/2026"
ws["C2"].font = subtitle_font
ws["E2"] = "Lineas:"
ws["E2"].font = Font(name="Arial", bold=True, size=10)
ws["F2"] = len(INVENTARIO)
ws["F2"].font = Font(name="Arial", bold=True, size=12, color="1E3A5F")
ws["G2"] = "Unidades (suma Cantidad):"
ws["G2"].font = Font(name="Arial", bold=True, size=10)
ws["H2"] = sum(r[6] for r in INVENTARIO)
ws["H2"].font = Font(name="Arial", bold=True, size=12, color="1E3A5F")
ws["A3"] = (
    "Codigo = ID por LINEA de inventario (no por unidad). Formato MARCA-FAMILIA-###. "
    "Cantidad = unidades en esa linea. Enlace = ficha o busqueda oficial."
)
ws["A3"].font = Font(name="Arial", size=9, color="6B7280", italic=True)
ws.merge_cells("A3:K3")

headers = [
    "#", "Codigo", "Marca", "Tipo", "Categoria", "Herramienta",
    "Modelo", "Cantidad", "Estado", "Enlace marca", "URL",
]
for col, h in enumerate(headers, 1):
    cell = ws.cell(4, col, h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center
    cell.border = thin

for i, row in enumerate(INVENTARIO, 1):
    r = i + 4
    codigo, marca, tipo, cat, nombre, modelo, cantidad, estado, url = row
    values = [i, codigo, marca, tipo, cat, nombre, modelo, cantidad, estado]
    for col, val in enumerate(values, 1):
        cell = ws.cell(r, col, val)
        cell.font = body_font
        cell.border = thin
        cell.alignment = center if col in (1, 2, 3, 4, 7, 8, 9) else left
        if i % 2 == 0 and col not in (2, 8, 9):
            cell.fill = alt_fill
    # Codigo resaltado
    ws.cell(r, 2).fill = code_fill
    ws.cell(r, 2).font = Font(name="Consolas", size=10, bold=True)
    # Estado
    est = ws.cell(r, 9)
    est.fill = comprado_fill if estado == "Comprado" else tengo_fill

    link_cell = ws.cell(r, 10, f"Ver en {marca}")
    link_cell.font = link_font
    link_cell.alignment = center
    link_cell.border = thin
    link_cell.hyperlink = url

    url_cell = ws.cell(r, 11, url)
    url_cell.font = Font(name="Arial", size=8, color="6B7280")
    url_cell.alignment = left
    url_cell.border = thin
    url_cell.hyperlink = url

last_inv = 4 + len(INVENTARIO)
dv = DataValidation(type="list", formula1='"Comprado,Ya tengo"', allow_blank=False)
ws.add_data_validation(dv)
dv.add(f"I5:I{last_inv}")

for i, w in enumerate([4, 16, 12, 12, 14, 40, 14, 10, 11, 16, 50], 1):
    ws.column_dimensions[get_column_letter(i)].width = w
ws.auto_filter.ref = f"A4:K{last_inv}"
ws.freeze_panes = "A5"
ws.row_dimensions[1].height = 24
ws.row_dimensions[4].height = 22

# ===== Pendientes =====
wp = wb.create_sheet("Pendientes")
wp["A1"] = "Pendientes de compra - Arsenal Multioficio 2026"
wp["A1"].font = title_font
wp.merge_cells("A1:I1")
wp["A2"] = (
    "Prioridad 1 = mas urgente. Codigo = ID reservado (misma logica por linea). "
    "Montos en colones (CRC)."
)
wp["A2"].font = subtitle_font

ph = [
    "Prioridad", "Codigo", "Marca", "Herramienta", "Modelo",
    "Precio (CRC)", "Notas", "Enlace marca", "URL",
]
for col, h in enumerate(ph, 1):
    cell = wp.cell(4, col, h)
    cell.font = header_font
    cell.fill = PatternFill("solid", fgColor="B45309")
    cell.alignment = center
    cell.border = thin

for i, row in enumerate(PENDIENTES, 1):
    r = i + 4
    codigo, prio, marca, nombre, modelo, precio, notas, url = row
    for col, val in enumerate([prio, codigo, marca, nombre, modelo, precio, notas], 1):
        cell = wp.cell(r, col, val)
        cell.font = body_font
        cell.border = thin
        cell.alignment = center if col in (1, 2, 3, 5) else left
        cell.fill = pend_fill
        if col == 6 and val is not None:
            cell.number_format = '"₡"#,##0'
    wp.cell(r, 2).fill = code_fill
    wp.cell(r, 2).font = Font(name="Consolas", size=10, bold=True)
    lc = wp.cell(r, 8, f"Ver en {marca}")
    lc.font = link_font
    lc.alignment = center
    lc.border = thin
    lc.fill = pend_fill
    lc.hyperlink = url
    uc = wp.cell(r, 9, url)
    uc.font = Font(name="Arial", size=8, color="6B7280")
    uc.border = thin
    uc.fill = pend_fill
    uc.hyperlink = url

last_p = 4 + len(PENDIENTES)
wp.cell(last_p + 2, 5, "Total estimado (con precio):")
wp.cell(last_p + 2, 5).font = Font(name="Arial", bold=True, size=10)
wp.cell(last_p + 2, 5).alignment = Alignment(horizontal="right")
wp.cell(last_p + 2, 6, f"=SUM(F5:F{last_p})")
wp.cell(last_p + 2, 6).font = Font(name="Arial", bold=True, size=11, color="B45309")
wp.cell(last_p + 2, 6).number_format = '"₡"#,##0'
wp.cell(last_p + 2, 6).fill = pend_fill

for i, w in enumerate([10, 14, 12, 40, 14, 14, 14, 16, 50], 1):
    wp.column_dimensions[get_column_letter(i)].width = w
wp.freeze_panes = "A5"
wp.auto_filter.ref = f"A4:I{last_p}"

# ===== Consumibles =====
wc = wb.create_sheet("Consumibles")
wc["A1"] = "Consumibles - Arsenal Multioficio 2026"
wc["A1"].font = title_font
wc.merge_cells("A1:L1")
wc["A2"] = (
    "Clase = familia. Tipo terminal si es conector. Ubicacion = codigo del contenedor "
    "(ej. MW-PKO-005 = PACKOUT Organizer 19). Por COLOR. NO pieza a pieza."
)
wc["A2"].font = subtitle_font
wc.merge_cells("A2:L2")
wc["A3"] = (
    "Codigo CNS-*-###. Clases: " + " | ".join(CLASES_CONSUMIBLE)
    + f". Ubic. PACKOUT actual conectores: {UBIC_PKO_ORG19}. Clic en Nombre = link compra."
)
wc["A3"].font = Font(name="Arial", size=9, color="6B7280", italic=True)
wc.merge_cells("A3:L3")

ch = [
    "#", "Codigo", "Clase", "Tipo terminal", "Nombre / kit", "Marca / ref",
    "Forma", "Cantidad", "Umbral reorden", "Estado", "Ubicacion", "Notas / URL",
]
for col, h in enumerate(ch, 1):
    cell = wc.cell(4, col, h)
    cell.font = header_font
    cell.fill = cns_header
    cell.alignment = center
    cell.border = thin

# Relleno visual por color de conector (ROJO / AZUL / AMARILLO)
color_row_fills = {
    "ROJO": PatternFill("solid", fgColor="FECACA"),
    "AZUL": PatternFill("solid", fgColor="BFDBFE"),
    "AMARILLO": PatternFill("solid", fgColor="FDE68A"),
    "BLANCO": PatternFill("solid", fgColor="F3F4F6"),
    "WHITE": PatternFill("solid", fgColor="F3F4F6"),
    "ROSA": PatternFill("solid", fgColor="FBCFE8"),
    "PINK": PatternFill("solid", fgColor="FBCFE8"),
}
ubic_fill = PatternFill("solid", fgColor="DDD6FE")  # violeta suave para ubicacion

for i, row in enumerate(CONSUMIBLES, 1):
    r = i + 4
    codigo, clase, tipo_term, nombre, marca_ref, forma, cantidad, umbral, estado, ubicacion, notas, url = row
    row_fill = cns_fill
    up = (nombre or "").upper()
    for key, fill in color_row_fills.items():
        if key in up:
            row_fill = fill
            break
    vals = [i, codigo, clase, tipo_term, nombre, marca_ref, forma, cantidad, umbral, estado, ubicacion]
    for col, val in enumerate(vals, 1):
        cell = wc.cell(r, col, val)
        cell.font = body_font
        cell.border = thin
        cell.alignment = center if col in (1, 2, 7, 8, 10, 11) else left
        cell.fill = row_fill
    wc.cell(r, 2).fill = code_fill
    wc.cell(r, 2).font = Font(name="Consolas", size=10, bold=True)
    wc.cell(r, 3).font = Font(name="Arial", size=10, bold=True)  # Clase
    wc.cell(r, 4).font = Font(name="Arial", size=10, bold=True)  # Tipo terminal
    # Ubicacion = codigo de lugar
    ucell = wc.cell(r, 11)
    ucell.fill = ubic_fill
    ucell.font = Font(name="Consolas", size=10, bold=True)
    est = wc.cell(r, 10)
    est.fill = reponer_fill if estado == "Reponer" else tengo_cns_fill
    note_cell = wc.cell(r, 12, notas)
    note_cell.font = body_font
    note_cell.border = thin
    note_cell.alignment = left
    note_cell.fill = row_fill
    name_cell = wc.cell(r, 5)
    name_cell.hyperlink = url
    name_cell.font = link_font

last_c = 4 + len(CONSUMIBLES)
dv_c = DataValidation(type="list", formula1='"Tengo,Reponer"', allow_blank=False)
wc.add_data_validation(dv_c)
dv_c.add(f"J5:J{last_c}")
dv_f = DataValidation(type="list", formula1='"kit,paquete,aprox"', allow_blank=False)
wc.add_data_validation(dv_f)
dv_f.add(f"G5:G{last_c}")
dv_clase = DataValidation(
    type="list",
    formula1='"' + ",".join(CLASES_CONSUMIBLE) + '"',
    allow_blank=False,
)
wc.add_data_validation(dv_clase)
dv_clase.add(f"C5:C{last_c}")
dv_t = DataValidation(
    type="list",
    formula1='"Spade (pala / quick disconnect),Butt (empalme / solder seal),Butt (empalme / heat shrink crimp),Butt (empalme / non-insulated crimp),Ring (anillo),Fork (horquilla),Wire nut,—,Otro"',
    allow_blank=True,
)
wc.add_data_validation(dv_t)
dv_t.add(f"D5:D{last_c}")
# Ubicaciones PACKOUT conocidas
dv_u = DataValidation(
    type="list",
    formula1='"MW-PKO-001,MW-PKO-002,MW-PKO-005,MW-PKO-003,MW-PKO-004,Taller,Vehiculo,Otro"',
    allow_blank=True,
)
wc.add_data_validation(dv_u)
dv_u.add(f"K5:K{last_c}")

# fila ayuda
help_r = last_c + 2
wc.cell(help_r, 1, "Como usar:").font = Font(name="Arial", bold=True, size=10)
wc.cell(help_r + 1, 1, "1) Clase = familia (Conector electrico, Cinta / aislante, Fijacion...).").font = Font(name="Arial", size=9, color="4B5563")
wc.cell(help_r + 2, 1, "2) Tipo terminal solo si Clase=Conector electrico; si no, '—'.").font = Font(name="Arial", size=9, color="4B5563")
wc.cell(help_r + 3, 1, "3) Ubicacion = codigo del contenedor (MW-PKO-005 = Organizer 19 pulg PACKOUT).").font = Font(name="Arial", size=9, color="4B5563")
wc.cell(help_r + 4, 1, "4) Por COLOR en filas. Reponer si baja umbral. Clic en Nombre (col E) = link.").font = Font(name="Arial", size=9, color="4B5563")

for i, w in enumerate([4, 16, 18, 28, 40, 20, 10, 10, 12, 10, 14, 42], 1):
    wc.column_dimensions[get_column_letter(i)].width = w
wc.freeze_panes = "A5"
wc.auto_filter.ref = f"A4:L{last_c}"
wc.row_dimensions[1].height = 24
wc.row_dimensions[4].height = 22

# ===== Resumen =====
wr = wb.create_sheet("Resumen", 0)
wr["A1"] = "Arsenal Multioficio 2026 - Resumen"
wr["A1"].font = title_font
wr.merge_cells("A1:D1")
wr["A2"] = "Saul Marin Gonzalez | Codigo por linea de inventario | Respaldo local"
wr["A2"].font = subtitle_font

wr["A4"] = "Metricas del inventario"
wr["A4"].font = Font(name="Arial", bold=True, size=12, color="1E3A5F")

for cell_ref, label in [
    ("A5", "Lineas de inventario (codigos)"),
    ("A6", "Unidades totales (suma Cantidad)"),
    ("A7", "Estado: Ya tengo (lineas)"),
    ("A8", "Estado: Comprado (lineas)"),
    ("A9", "Pendientes de compra"),
    ("A10", "Costo estimado pendientes (CRC)"),
    ("A11", "Lineas de consumibles (kits)"),
    ("A12", "Consumibles a reponer"),
]:
    wr[cell_ref] = label
    wr[cell_ref].font = body_font
    wr[cell_ref].fill = resumen_fill
    wr[cell_ref].border = thin

wr["B5"] = len(INVENTARIO)
wr["B6"] = sum(r[6] for r in INVENTARIO)
wr["B7"] = f'=COUNTIF(Inventario!I5:I{last_inv},"Ya tengo")'
wr["B8"] = f'=COUNTIF(Inventario!I5:I{last_inv},"Comprado")'
wr["B9"] = len(PENDIENTES)
wr["B10"] = f"=SUM(Pendientes!F5:F{last_p})"
wr["B10"].number_format = '"₡"#,##0'
wr["B11"] = len(CONSUMIBLES)
wr["B12"] = f'=COUNTIF(Consumibles!J5:J{last_c},"Reponer")'
for r in range(5, 13):
    wr.cell(r, 2).font = Font(name="Arial", bold=True, size=12)
    wr.cell(r, 2).border = thin
    wr.cell(r, 2).alignment = center

wr["A14"] = "Esquema de codigos (por LINEA)"
wr["A14"].font = Font(name="Arial", bold=True, size=12, color="1E3A5F")
scheme = [
    "Inventario: MARCA-FAMILIA-###  (ej. MW-M12-001, KL-ELC-003, MW-M12-BAT-001)",
    "Un codigo = una linea, no una unidad fisica. Cantidad en columna aparte.",
    "El ### no se reutiliza. Pendientes: codigo reservado hasta comprar.",
    "Consumibles: CNS-*-### + CLASE + tipo terminal + color + UBICACION (codigo del lugar).",
    "Clases: Conector electrico | Cinta / aislante | Fijacion | Corte / desgaste | Quimico / sellado | Otro",
    "Ubicacion ejemplo: MW-PKO-005 = PACKOUT Organizer 19. Prefijos MW/KL/CNS... NO pieza a pieza.",
]
for i, t in enumerate(scheme):
    wr.cell(15 + i, 1, t).font = Font(name="Arial", size=9, color="374151")

wr["A22"] = "Sitios oficiales de marca"
wr["A22"].font = Font(name="Arial", bold=True, size=12, color="1E3A5F")
brand_home = [
    ("Milwaukee", "https://www.milwaukeetool.com/"),
    ("Klein Tools", "https://www.kleintools.com/"),
    ("Knipex", "https://www.knipex.com/"),
    ("Wera", "https://www.wera.de/en"),
    ("Huepar", "https://huepar.com/"),
    ("LEXIVON", "https://www.lexivon.com/"),
    ("Skil", "https://www.skil.com/"),
]
wr["A23"] = "Marca"
wr["B23"] = "Sitio web"
for col in (1, 2):
    wr.cell(23, col).font = header_font
    wr.cell(23, col).fill = header_fill
    wr.cell(23, col).border = thin
for i, (name, url) in enumerate(brand_home):
    r = 24 + i
    wr.cell(r, 1, name).font = body_font
    wr.cell(r, 1).border = thin
    c = wr.cell(r, 2, url)
    c.font = link_font
    c.hyperlink = url
    c.border = thin

last_b = 23 + len(brand_home)
wr.cell(last_b + 2, 1, "Lineas por marca").font = Font(name="Arial", bold=True, size=12, color="1E3A5F")
hdr = last_b + 3
wr.cell(hdr, 1, "Marca").font = header_font
wr.cell(hdr, 1).fill = header_fill
wr.cell(hdr, 1).border = thin
wr.cell(hdr, 2, "Lineas").font = header_font
wr.cell(hdr, 2).fill = header_fill
wr.cell(hdr, 2).border = thin

marcas = sorted(set(r[1] for r in INVENTARIO))
for i, m in enumerate(marcas):
    r = hdr + 1 + i
    wr.cell(r, 1, m).font = body_font
    wr.cell(r, 1).border = thin
    wr.cell(r, 2, f"=COUNTIF(Inventario!C5:C{last_inv},A{r})")
    wr.cell(r, 2).font = body_font
    wr.cell(r, 2).border = thin
    wr.cell(r, 2).alignment = center

wr["D4"] = "Ecosistema"
wr["D4"].font = Font(name="Arial", bold=True, size=12, color="1E3A5F")
for i, t in enumerate(
    [
        "Base principal: Milwaukee M12 + M18",
        "Prueba / electrico: Klein Tools",
        "Alicates y llaves: Knipex + Wera Joker",
        "Organizacion: PACKOUT",
        "ID interno: columna Codigo (por linea)",
        "Consumibles: pestaña aparte (kits CNS-*)",
        "Enlaces: Inventario col J/K; Consumibles clic en nombre",
    ]
):
    wr.cell(5 + i, 4, t).font = body_font
    wr.cell(5 + i, 4).fill = resumen_fill
    wr.cell(5 + i, 4).border = thin

wr.column_dimensions["A"].width = 42
wr.column_dimensions["B"].width = 48
wr.column_dimensions["C"].width = 3
wr.column_dimensions["D"].width = 55

out = Path(__file__).resolve().parent.parent / "inventario" / "Arsenal_Multioficio_2026.xlsx"
out.parent.mkdir(parents=True, exist_ok=True)
try:
    wb.save(out)
    print("OK", out)
except PermissionError:
    alt = out.with_name("Arsenal_Multioficio_2026_new.xlsx")
    wb.save(alt)
    print("LOCKED original -> saved", alt)
    print("Cerra el Excel y volve a correr el script, o renombra _new.")
print("Lineas inventario:", len(INVENTARIO), "| Unidades:", sum(r[6] for r in INVENTARIO))
print("Pendientes:", len(PENDIENTES))
print("Consumibles:", len(CONSUMIBLES))
codes = [r[0] for r in INVENTARIO] + [r[0] for r in PENDIENTES] + [r[0] for r in CONSUMIBLES]
assert len(codes) == len(set(codes)), "Codigos duplicados!"
print("Codigos unicos: OK")
